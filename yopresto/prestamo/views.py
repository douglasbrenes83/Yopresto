from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
# Create your views here.
# prestamos/views.py
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from .models import Pago,Prestamo,Cliente
from .forms import PrestamoForm,PagoForm
from django.utils.dateparse import parse_date
from datetime import datetime
from django.http import JsonResponse
from django.db.models import Sum

class PrestamoListView(LoginRequiredMixin,ListView):
    login_url = 'index' 
    model = Prestamo
    template_name = 'prestamos/prestamo_list.html'
    context_object_name = 'prestamos'

class PrestamoCreateView(LoginRequiredMixin,CreateView):
    login_url = 'index' 
    model = Prestamo
    form_class = PrestamoForm
    template_name = 'prestamos/prestamo_form.html'
    success_url = reverse_lazy('prestamo_list')

class PrestamoUpdateView(LoginRequiredMixin,UpdateView):
    login_url = 'index' 
    model = Prestamo
    form_class = PrestamoForm
    template_name = 'prestamos/editar_prestamo.html'
    success_url = reverse_lazy('prestamo_list')

    def form_valid(self, form):
        monto = form.cleaned_data['monto']
        tasa_interes = form.cleaned_data['tasa_interes']
        # Recalcular saldo_actual
        form.instance.saldo_actual = monto + (monto * tasa_interes / 100)
        return super().form_valid(form)


class PrestamoDeleteView(LoginRequiredMixin,DeleteView):
    login_url = 'index'
    model = Prestamo
    template_name = 'prestamos/prestamo_delete.html'
    success_url = reverse_lazy('prestamo_list')

#pagos

class PrestamoPagosDetailView(LoginRequiredMixin,DetailView):
    login_url = 'index'
    model = Prestamo
    template_name = 'prestamos/detalle_prestamo.html'
    context_object_name = 'prestamo'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        prestamo = self.object
        context['pagos'] = prestamo.pagos.all()  # Obtener todos los pagos relacionados con este préstamo
        return context
    
class PagoCreateView(LoginRequiredMixin,CreateView):
    login_url = 'index'
    model = Pago
    form_class = PagoForm
    template_name = 'pagos/pago_create.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        prestamo = get_object_or_404(Prestamo, id=self.kwargs['prestamo_id'])
        kwargs['prestamo'] = prestamo  # Pasar el préstamo al formulario
        return kwargs

    def form_valid(self, form):
        prestamo = get_object_or_404(Prestamo, id=self.kwargs['prestamo_id'])
        form.instance.prestamo = prestamo
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        prestamo = get_object_or_404(Prestamo, id=self.kwargs['prestamo_id'])
        context['prestamo'] = prestamo
        return context

    def get_success_url(self):
        return reverse_lazy('prestamo_pagos', kwargs={'pk': self.object.prestamo.id})

class PrestamoActivoListView(LoginRequiredMixin,ListView):
    login_url = 'index'
    model = Prestamo
    template_name = 'prestamos/prestamo_activos.html'
    context_object_name = 'prestamos'

    def get_queryset(self):
        # Filtra solo los préstamos activos
        return Prestamo.objects.filter(estado="Activo")
    
def pagos_por_rango(request):
    pagos = []
    total_pagado = 0
    fecha_inicio = None
    fecha_fin = None

    if request.method == "GET":
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        if fecha_inicio and fecha_fin:
            try:
                # Convertir las fechas de string a objetos de fecha
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

                # Filtrar pagos dentro del rango
                pagos = Pago.objects.filter(fecha_pago__range=(fecha_inicio, fecha_fin)).select_related('prestamo')

                # Sumar el total de los pagos en ese rango de fechas
                total_pagado = pagos.aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0
            except ValueError:
                # Manejar error si las fechas tienen formato incorrecto
                return render(request, 'error.html', {'mensaje': 'Fechas no válidas'})

    # Preparar el contexto para la plantilla
    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'pagos': pagos,
        'total_pagado': total_pagado,
    }
   
    return render(request, 'prestamos/pagos_rango.html', context)
def nuevo_prestamo(request, cliente_id=None):
    initial_data = {}
    if cliente_id:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        initial_data['cliente'] = cliente

    if request.method == 'POST':
        form = PrestamoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('prestamo_list')
    else:
        form = PrestamoForm(initial=initial_data)

    return render(request, 'prestamos/nuevo_prestamo.html', {'form': form})


#muestra un reporte de prestamos 
import datetime
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .models import Prestamo

def reporte_prestamos(request):
    # Obtener la fecha actual
    hoy = datetime.date.today()
    primer_dia_mes = datetime.date(hoy.year, hoy.month, 1)
    ultimo_dia_mes = datetime.date(hoy.year, hoy.month + 1, 1) - datetime.timedelta(days=1)

    # Filtrar préstamos del mes actual
    prestamos = Prestamo.objects.filter(fecha_inicio__range=[primer_dia_mes, ultimo_dia_mes])

    # Preparar datos para la vista
    total_prestamos = 0
    total_intereses = 0
    datos = []

    for prestamo in prestamos:
        interes = prestamo.monto * (prestamo.tasa_interes / 100)
        total_prestamos += prestamo.monto
        total_intereses += interes

        datos.append({
            'cliente': prestamo.cliente.nombre,
            'cedula': prestamo.cliente.cedula,
            'monto': prestamo.monto,
            'tasa_interes': prestamo.tasa_interes,
            'interes_mensual': round(interes, 2),
            'fecha_inicio': prestamo.fecha_inicio,
            'fecha_vencimiento': prestamo.fecha_vencimiento,
            'saldo_actual': prestamo.saldo_actual,
            'estado': prestamo.estado,
        })

    contexto = {
        'datos': datos,
        'total_prestamos': total_prestamos,
        'total_intereses': total_intereses,
        'mes': hoy.strftime('%B %Y')
    }

    return render(request, 'reporte_prestamos.html', contexto)

def descargar_reporte_excel(request):
    hoy = datetime.date.today()
    primer_dia_mes = datetime.date(hoy.year, hoy.month, 1)
    ultimo_dia_mes = datetime.date(hoy.year, hoy.month + 1, 1) - datetime.timedelta(days=1)

    prestamos = Prestamo.objects.filter(fecha_inicio__range=[primer_dia_mes, ultimo_dia_mes])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {hoy.strftime('%B %Y')}"

    headers = ['Cliente', 'Cédula', 'Monto', 'Tasa de Interés (%)', 'Interés Mensual', 'Fecha de Inicio', 'Fecha de Vencimiento', 'Saldo Actual', 'Estado']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    for row_num, prestamo in enumerate(prestamos, 2):
        interes = prestamo.monto * (prestamo.tasa_interes / 100)
        ws.append([
            prestamo.cliente.nombre,
            prestamo.cliente.cedula,
            float(prestamo.monto),
            float(prestamo.tasa_interes),
            round(interes, 2),
            prestamo.fecha_inicio,
            prestamo.fecha_vencimiento,
            float(prestamo.saldo_actual),
            prestamo.estado
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_prestamos_{hoy.strftime("%Y_%m")} .xlsx"'
    wb.save(response)

    return response
#reporte de prestamos vencidos
from datetime import date
from django.shortcuts import render
from .models import Prestamo
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

def prestamos_vencidos(request):
    hoy = date.today()
    prestamos = Prestamo.objects.filter(fecha_vencimiento__lt=hoy)

    for prestamo in prestamos:
        prestamo.dias_vencidos = (hoy - prestamo.fecha_vencimiento).days

    return render(request, 'prestamos/prestamos_vencidos.html', {'prestamos': prestamos})

def exportar_prestamos_vencidos(request):
    hoy = date.today()
    prestamos = Prestamo.objects.filter(fecha_vencimiento__lt=hoy)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Préstamos Vencidos'

    # Encabezados
    headers = ['ID', 'Cliente', 'Monto', 'Fecha de Vencimiento', 'Días Vencidos']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Datos
    for row, prestamo in enumerate(prestamos, 2):
        dias_vencidos = (hoy - prestamo.fecha_vencimiento).days
        ws.cell(row=row, column=1, value=prestamo.id)
        ws.cell(row=row, column=2, value=str(prestamo.cliente))
        ws.cell(row=row, column=3, value=prestamo.monto)
        ws.cell(row=row, column=4, value=prestamo.fecha_vencimiento.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=dias_vencidos)

    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

    # Descargar archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=prestamos_vencidos.xlsx'
    wb.save(response)
    return response
